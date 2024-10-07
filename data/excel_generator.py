# data/excel_generator.py
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import os
import calendar
from datetime import datetime

class ExcelGenerator:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "製造工程"

    def gregorian_to_reiwa(self, year, month):
        """
        西暦を令和に変換する関数。
        令和元年は2019年。
        """
        if year < 2019:
            raise ValueError("令和は2019年以降の年です。")
        reiwa_year = year - 2018
        return f"令和{reiwa_year}年{month}月"
    
    def font_template(self, start_row, end_row):
        """
        表の入力部分の行高さ、フォントを設定。
        generate_excel内でfor文により各表に適用。
        """
        start_col = 1 # A列
        end_col = 36 # AJ列
        # ヘッダー部分以外
        for row in range(start_row + 3, end_row):
            self.ws.row_dimensions[row].height = 36.8
            for col in range(start_col, end_col + 1):
                cell = self.ws.cell(row=row, column=col)
        # ヘッダー部分含む
        for row in range(start_row, end_row):
            for col in range(start_col, end_col + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.font = Font(name='游ゴシック', size=14, bold=True, color='000000') # フォントの設定
                cell.alignment = Alignment(horizontal="center", vertical="center") # 文字位置の設定

        
    
    def is_holiday(self, year, month, day):
        """
        各工場の休日判定。
        休日セルの色付けに使用。
        """
        # 年と月を文字列に変換
        year_str = str(year)
        month_str = str(month)
        
        # 工場ごとの休日データを取得
        if not hasattr(self, 'holidays_data'):
            raise AttributeError("holidays_dataがロードされていません。generate_excelを先に実行してください。")
        
        # 指定された年月の休日リストを取得
        if year_str in self.holidays_data and month_str in self.holidays_data[year_str]:
            return day in self.holidays_data[year_str][month_str]
        return False

    def save_and_open(self, output_path):
        """
        生成したexcelの一時保存と表示。
        """
        self.wb.save(output_path)
        os.startfile(output_path)

    def generate_excel(self, title, start_date, end_date, factory, output_path):
        """
        excelの生成。
        """
        try:
            # 休日データのロード
            holidays_file = os.path.join(os.path.dirname(__file__), 'holidays', f'{factory}.json')
            if not os.path.exists(holidays_file):
                raise FileNotFoundError(f"工場 '{factory}' の休日JSONファイルが見つかりません。")
            
            with open(holidays_file, 'r', encoding='utf-8') as f:
                self.holidays_data = json.load(f)
            
            # セル幅の設定
            self.ws.sheet_view.zoomScale = 55
            self.ws.column_dimensions['A'].width = 25
            self.ws.column_dimensions['B'].width = 30
            for col in range(3, 38): # C列~AL列まで
                self.ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 6

            # 作成日を入力
            now = datetime.now()
            current_date_str = now.strftime('%Y/%m/%d')
            self.ws.merge_cells(start_row=1, start_column=31, end_row=1, end_column=36) # セル結合(AE1:AJ1)
            self.ws.row_dimensions[1].height = 32.5
            self.ws['AE1'].value = current_date_str
            self.ws['AE1'].font = Font(name='游ゴシック', size=20, bold=True, color='000000')
            self.ws['AE1'].alignment = Alignment(horizontal="right")

            # タイトルを入力
            self.ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=36) # セル結合(A3:AJ3)
            self.ws.row_dimensions[3].height = 58.5
            self.ws['A3'].value = f"{title} 製造工程計画"
            self.ws['A3'].font = Font(name='游ゴシック', size=36, bold=True, color='000000')
            self.ws['A3'].alignment = Alignment(horizontal="center")

            # 定型コメントを入力
            self.ws.row_dimensions[4].height = 32.5
            self.ws['A4'].value = "※以下の日程は生コン打設となります。"
            self.ws['A4'].font = Font(name='游ゴシック', size=20, bold=True, color='000000')
            self.ws['AJ4'].value = "ベルテクス株式会社"
            self.ws['AJ4'].font = Font(name='游ゴシック', size=20, bold=True, color='000000')
            self.ws['AJ4'].alignment = Alignment(horizontal="right")

            # テーブルの開始位置
            first_table_start_row = 9
            contents_rows = 8  # 項目の数
            rows_per_table = contents_rows + 5

            # 現在の行を初期化
            current_row = first_table_start_row

            # 開始年月から終了年月までループ
            year = start_date.year
            month = start_date.month
            end_year = end_date.year
            end_month = end_date.month
            
            # 各月の表ループ
            while (year < end_year) or (year == end_year and month <= end_month):
                # 各表のヘッダー部分作成
                self.ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=1) # セル結合(項目)
                items_cell = self.ws.cell(row=current_row, column=1, value="項目")
                self.ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + 2, end_column=2) # セル結合(規格)
                standards_cell = self.ws.cell(row=current_row, column=2, value="規格")
                self.ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row + 2, end_column=3) # セル結合(受注)
                orders_cell = self.ws.cell(row=current_row, column=3, value="受注")
                self.ws.merge_cells(start_row=current_row, start_column=35, end_row=current_row + 1, end_column=36) # セル結合(合計)
                sums_cell = self.ws.cell(row=current_row, column=35, value="合計")
                volumes_cell = self.ws.cell(row=current_row + 2, column=35, value="数量")
                remains_cell = self.ws.cell(row=current_row + 2, column=36, value="残り")

                # 初期値、数式の設定
                for row_offset in range(3, contents_rows + 3):
                    cell_row = current_row + row_offset
                    self.ws.cell(row=cell_row, column=3, value="0") # 受注初期値＝０
                    self.ws.cell(row=cell_row, column=35, value=f"=SUM(D{cell_row}:AH{cell_row})") # 合計数量の算出
                for row_offset in range(3, contents_rows + 3): # 残り数量の算出
                    cell_row = current_row + row_offset
                    if current_row == first_table_start_row: # 1つ目とそれ以降で計算を変更
                        self.ws.cell(row=cell_row, column=36, value=f"=C{str(cell_row)}-AI{str(cell_row)}")
                    else:
                        self.ws.cell(row=cell_row, column=36, value=f"=AJ{str(cell_row - rows_per_table)}-AI{str(cell_row)}")
                
                # 全体に点線
                bold_side = Side(style='medium', color='000000') # 太線
                thin_side = Side(style='thin', color='000000') # 細線
                dashed_side = Side(style='hair', color='000000') # 薄い点線
                for row in range(current_row + 3, current_row + rows_per_table - 2):
                    for col in range(4, 35):
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=dashed_side, bottom=dashed_side, right=dashed_side, left=dashed_side)
                    for col in [1]:
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=dashed_side, bottom=dashed_side, right=thin_side, left=bold_side)
                    for col in [2,3]:
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=dashed_side, bottom=dashed_side, right=thin_side, left=thin_side)
                    for col in [35]:
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=dashed_side, bottom=dashed_side, right=thin_side, left=bold_side)
                    for col in [36]:
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=dashed_side, bottom=dashed_side, right=bold_side, left=thin_side)
                for row in [current_row]:
                    for col in range(4, 35):
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=bold_side, bottom=thin_side)
                for row in [current_row + 1]:
                    for col in range(4, 35):
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=thin_side, bottom=dashed_side, right=dashed_side, left=dashed_side)
                for row in [current_row + 2]:
                    for col in range(4, 35):
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=dashed_side, bottom=bold_side, right=dashed_side, left=dashed_side)
                for row in range(current_row, current_row + 3):
                    for col in [1, 2, 3]:
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=bold_side, bottom=bold_side, right=bold_side, left=bold_side)
                for row in range(current_row, current_row + 2):
                    for col in [35, 36]:
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=bold_side, bottom=thin_side, right=bold_side, left=bold_side)
                for row in [current_row + 2]:
                    for col in [35]:
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=thin_side, bottom=bold_side, right=thin_side, left=bold_side)
                    for col in [36]:
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=thin_side, bottom=bold_side, right=bold_side, left=thin_side)
                for row in [current_row + rows_per_table - 2]:
                    for col in range(1, 37):
                        cell = self.ws.cell(row=row, column=col)
                        cell.border = Border(top=bold_side)


                # 月のタイトルを入力
                reiwa_month = self.gregorian_to_reiwa(year, month)
                self.ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=34) # セル結合
                self.ws.row_dimensions[current_row].height = 32.5
                title_cell = self.ws.cell(row=current_row, column=4, value=reiwa_month)

                # 月の日数を取得
                last_day = calendar.monthrange(year, month)[1]

                # 日にちを入力
                self.ws.row_dimensions[current_row + 1].height = 23
                for day in range(1, last_day + 1):
                    col = 4 + day - 1  # D列は4
                    cell = self.ws.cell(row=current_row + 1, column=col, value=day)
                
                # 曜日を入力
                self.ws.row_dimensions[current_row + 2].height = 23
                for day in range(1, last_day + 1):
                    date = datetime(year, month, day)
                    weekday_en = date.strftime('%a')  # 'Mon', 'Tue', etc.
                    # 英語の曜日を日本語にマッピング
                    weekday_jp = {
                        'Mon': '月',
                        'Tue': '火',
                        'Wed': '水',
                        'Thu': '木',
                        'Fri': '金',
                        'Sat': '土',
                        'Sun': '日'
                    }.get(weekday_en, '')
                    col = 4 + day - 1
                    self.ws.cell(row=current_row + 2, column=col, value=weekday_jp)

                # フォントの適用
                self.font_template(current_row, current_row + rows_per_table - 2)
                title_cell.font = Font(name='游ゴシック', size=20, bold=True, color='000000')

                # 休日のセルに色を付ける（楯列すべて）
                for day in range(1, last_day + 1):
                    if self.is_holiday(year, month, day):
                        col = 4 + day - 1
                        for row_offset in range(1, rows_per_table - 2):
                            cell_row = current_row + row_offset
                            cell = self.ws.cell(row=cell_row, column=col)
                            cell.font = Font(name='游ゴシック', size=14, bold=True, color='ff0000')
                            cell.fill = PatternFill(fill_type='solid', fgColor='ffc7ce')

                # 次の月のテーブル開始行を更新
                current_row += rows_per_table

                # 月をインクリメント
                if month == 12:
                    year += 1
                    month = 1
                else:
                    month += 1

        except Exception as e:
            messagebox.showerror("エラー", f"製造工程表の生成中にエラーが発生しました。\n{e}")

        # 保存して開く
        self.save_and_open(output_path)
