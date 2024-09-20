# excel_generator.py

import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar
from tkinter import messagebox
import config

def gregorian_to_reiwa(year, month):
    """
    西暦を令和に変換する関数。
    令和元年は2019年。
    """
    if year < 2019:
        raise ValueError("令和は2019年以降の年です。")
    reiwa_year = year - 2018
    return f"令和{reiwa_year}年{month}月"

def is_holiday(year, month, day):
    """
    指定された年、月、日が休日かどうかを判定します。
    """
    year_str = str(year)
    month_str = str(month)
    return day in config.HOLIDAYS.get(year_str, {}).get(month_str, [])

def generate_schedule(title, start_year, start_month, end_year, end_month):
    try:
        
        template_path = config.TEMPLATE_PATH
        if not os.path.exists(template_path):
            messagebox.showerror("エラー", f"テンプレートファイル '{template_path}' が見つかりません。")
            return
        
        # 出力ディレクトリの作成
        if not os.path.exists(config.SAVE_DIR):
            os.makedirs(config.SAVE_DIR)
        
        wb = load_workbook(template_path)
        ws = wb.active  # アクティブなシートを選択
        
        # 作成日を入力
        now = datetime.now()
        current_date_str = now.strftime('%Y/%m/%d')
        ws['AE1'].value = current_date_str
        
        # タイトルを入力
        ws['A3'].value = f"{title} 製造工程計画"
        
        # テーブルの開始位置
        first_table_start_row = 8
        rows_per_table = 13  # 各月のテーブルが13行を使用
        
        # 現在の行を初期化
        current_row = first_table_start_row
        
        # 開始年月から終了年月までループ
        year = start_year
        month = start_month
        
        while (year < end_year) or (year == end_year and month <= end_month):
            # 令和形式に変換
            reiwa_month = gregorian_to_reiwa(year, month)
            
            # 月をD行に入力
            ws.cell(row=current_row, column=4, value=reiwa_month)
            
            # 月の日数を取得
            last_day = calendar.monthrange(year, month)[1]
            
            # 日にちをD行+1に入力
            for day in range(1, last_day + 1):
                col = 4 + day - 1  # D列は4
                ws.cell(row=current_row + 1, column=col, value=day)
            
            # 曜日をD行+2に入力
            for day in range(1, last_day + 1):
                date = datetime(year, month, day)
                weekday_en = date.strftime('%a')  # 'Mon', 'Tue', etc.
                # 英語の曜日を日本語にマッピング
                weekday_jp = {
                    'Mon': '月',
                    'Tue': '火',
                    'Wed': '水',
                    'Thu': '木',
                    'Fri': '金',
                    'Sat': '土',
                    'Sun': '日'
                }.get(weekday_en, '')
                col = 4 + day - 1
                ws.cell(row=current_row + 2, column=col, value=weekday_jp)
            
            # 記入欄 (11行目～18行目) は空白のまま
            
            # 休日のセルに色を付ける
            for day in range(1, last_day + 1):
                if is_holiday(year, month, day):
                    col = 4 + day - 1
                    for row_offset in [1, 2]:  # 1: 日にち, 2: 曜日
                        cell_row = current_row + row_offset
                        cell = ws.cell(row=cell_row, column=col)
                        cell.font = Font(name='游ゴシック', size=14, bold=True, color='FF0000')  # フォント設定
            
            # 空白行19,20は自動で空白のまま
            
            # 次の月のテーブル開始行を更新
            current_row += rows_per_table
            
            # 月をインクリメント
            if month == 12:
                year += 1
                month = 1
            else:
                month += 1
        
        # 保存ファイル名の作成
        formatted_date = now.strftime("%Y-%m-%d_%H-%M-%S")
        save_filename = f"{title}製造工程表_{formatted_date}.xlsx"
        save_path = os.path.join(config.SAVE_DIR, save_filename)
        wb.save(save_path)
        
        messagebox.showinfo("成功", f"製造工程表を '{save_path}' として保存しました。")
    except Exception as e:
        messagebox.showerror("エラー", f"製造工程表の生成中にエラーが発生しました。\n{e}")
